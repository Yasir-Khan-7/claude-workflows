#!/usr/bin/env python3
"""
Compile analysis results into a markdown report for the AI Opportunity Radar.
"""

import json
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# Project root: Claude Workflows (parent of ai-opportunity-radar)
_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
_TMP_DIR = _PROJECT_ROOT / ".tmp"


def _profile_summary_md(profile: dict) -> str:
    """Build profile summary section."""
    lines = []
    if profile.get("name"):
        lines.append(f"- **Name:** {profile['name']}")
    if profile.get("headline"):
        lines.append(f"- **Headline:** {profile['headline']}")
    if profile.get("skills"):
        skills = profile["skills"]
        if isinstance(skills, list):
            skills_str = ", ".join(skills[:15])
            if len(skills) > 15:
                skills_str += f" (+{len(skills) - 15} more)"
        else:
            skills_str = str(skills)
        lines.append(f"- **Key Skills:** {skills_str}")
    if not lines:
        return "_No profile data provided._"
    return "\n".join(lines)


def _opportunities_table_md(opportunities: list[dict]) -> str:
    """Build opportunities summary table."""
    if not opportunities:
        return "_No opportunities found._"

    lines = [
        "| Rank | Job Title | Company | Location | Match Score | Link |",
        "|------|-----------|---------|-----------|--------------|------|",
    ]
    for i, opp in enumerate(opportunities, 1):
        title = (opp.get("title") or opp.get("job_title") or "—").replace("|", "\\|")
        company = (opp.get("company") or opp.get("employer") or "—").replace("|", "\\|")
        location = (opp.get("location") or opp.get("job_location") or "—").replace("|", "\\|")
        score = opp.get("match_score", "—")
        url = opp.get("url") or opp.get("link") or ""
        link_md = f"[View]({url})" if url else "—"
        lines.append(f"| {i} | {title} | {company} | {location} | {score} | {link_md} |")

    return "\n".join(lines)


def _opportunity_detail_md(opp: dict, rank: int) -> str:
    """Build detailed section for one opportunity."""
    title = opp.get("title") or opp.get("job_title") or "Unknown"
    company = opp.get("company") or opp.get("employer") or "Unknown"
    location = opp.get("location") or opp.get("job_location") or ""
    score = opp.get("match_score", "—")
    url = opp.get("url") or opp.get("link") or ""
    description = opp.get("description") or opp.get("snippet") or opp.get("body") or ""
    match_explanation = opp.get("match_explanation", "")
    application_tips = opp.get("application_tips", "")
    contact_email = opp.get("contact_email", "")
    hiring_manager = opp.get("hiring_manager", "")

    lines = [
        f"### {rank}. {title} — {company}",
        "",
        f"**Location:** {location}  \n**Match Score:** {score}/100",
        "",
        "#### Description",
        "",
        description[:2000] + ("..." if len(description) > 2000 else "") if description else "_No description available._",
        "",
        "#### Match Analysis",
        "",
        match_explanation or "_No analysis available._",
        "",
        "#### Application Tips",
        "",
        application_tips or "_No tips available._",
        "",
    ]

    if contact_email or hiring_manager:
        lines.append("#### Contact Information")
        lines.append("")
        if contact_email:
            lines.append(f"- **Email:** {contact_email}")
        if hiring_manager:
            lines.append(f"- **Hiring Manager:** {hiring_manager}")
        lines.append("")

    if url:
        lines.append(f"**Direct Link:** [Apply here]({url})")
        lines.append("")

    lines.append("---")
    lines.append("")
    return "\n".join(lines)


def _summary_stats_md(opportunities: list[dict], profile: dict) -> str:
    """Build summary statistics section."""
    total = len(opportunities)
    scores = [o.get("match_score") for o in opportunities if isinstance(o.get("match_score"), (int, float))]
    avg_score = round(sum(scores) / len(scores), 1) if scores else 0
    top_matches = [o for o in opportunities if (o.get("match_score") or 0) >= 70][:5]

    lines = [
        "### Summary Statistics",
        "",
        f"- **Total opportunities found:** {total}",
        f"- **Average match score:** {avg_score}",
        f"- **High-match opportunities (≥70):** {len([o for o in opportunities if (o.get('match_score') or 0) >= 70])}",
        "",
    ]

    if top_matches and profile.get("skills"):
        skills = profile["skills"] if isinstance(profile["skills"], list) else [profile["skills"]]
        lines.append("**Top skills matched across high-score roles:**")
        lines.append("")
        # Simple heuristic: skills mentioned in top match descriptions
        matched_skills = set()
        for opp in top_matches:
            desc = (opp.get("description") or "") + (opp.get("match_explanation") or "")
            for s in skills[:10]:
                if s and s.lower() in desc.lower():
                    matched_skills.add(s)
        if matched_skills:
            lines.append(", ".join(sorted(matched_skills)[:10]))
        else:
            lines.append(", ".join(skills[:5]))
        lines.append("")

    return "\n".join(lines)


def _generate_excel(
    profile: dict, opportunities: list[dict], role: str, location: str, excel_path: Path
) -> str:
    """Generate a styled Excel workbook with opportunity data."""
    wb = Workbook()

    # --- Sheet 1: Opportunities Overview ---
    ws = wb.active
    ws.title = "Opportunities"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap_align = Alignment(vertical="top", wrap_text=True)

    title_cell = ws.cell(row=1, column=1, value="AI Opportunity Radar Report")
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="2F5496")
    ws.merge_cells("A1:I1")

    info_data = [
        ("Profile:", profile.get("name", "N/A")),
        ("Target Role:", role),
        ("Location:", location),
        ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Opportunities:", str(len(opportunities))),
    ]
    for i, (label, value) in enumerate(info_data, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=value).font = Font(size=10)

    header_row = len(info_data) + 3
    headers = [
        "Rank", "Job Title", "Company", "Location", "Match Score",
        "Match Analysis", "Application Tips", "Contact Email", "Link"
    ]
    col_widths = [6, 35, 25, 25, 13, 50, 50, 30, 45]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    med_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for i, opp in enumerate(opportunities, start=1):
        row = header_row + i
        score = opp.get("match_score", 0)
        if isinstance(score, (int, float)):
            score_val = int(score)
        else:
            score_val = 0

        if score_val >= 80:
            row_fill = high_fill
        elif score_val >= 60:
            row_fill = med_fill
        else:
            row_fill = low_fill

        emails = opp.get("emails", [])
        contact = opp.get("contact_email", "")
        if isinstance(emails, list) and emails:
            contact = ", ".join(emails)
        elif not contact and isinstance(emails, str):
            contact = emails

        row_data = [
            i,
            opp.get("title") or opp.get("job_title") or "Unknown",
            opp.get("company") or "Unknown",
            opp.get("location") or opp.get("job_location") or "",
            score_val,
            opp.get("match_explanation", ""),
            opp.get("application_tips", ""),
            contact,
            opp.get("url") or "",
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = wrap_align
            cell.fill = row_fill

    ws.auto_filter.ref = f"A{header_row}:I{header_row + len(opportunities)}"
    ws.freeze_panes = f"A{header_row + 1}"

    # --- Sheet 2: Profile Summary ---
    ws2 = wb.create_sheet("Profile")
    ws2.cell(row=1, column=1, value="Profile Summary").font = Font(bold=True, size=14, color="2F5496")
    ws2.merge_cells("A1:B1")

    profile_fields = [
        ("Name", profile.get("name", "")),
        ("Headline", profile.get("headline", "")),
        ("Current Role", profile.get("current_role", "")),
        ("Company", profile.get("company", "")),
        ("Location", profile.get("location", "")),
        ("Skills", ", ".join(profile.get("skills", [])) if isinstance(profile.get("skills"), list) else str(profile.get("skills", ""))),
        ("Experience", "\n".join(profile.get("experience", [])) if isinstance(profile.get("experience"), list) else str(profile.get("experience", ""))),
        ("Education", "\n".join(profile.get("education", [])) if isinstance(profile.get("education"), list) else str(profile.get("education", ""))),
    ]
    for i, (label, value) in enumerate(profile_fields, start=3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=value).alignment = Alignment(wrap_text=True)

    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 80

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(excel_path))
    logger.info("Excel report saved to %s", excel_path)
    return str(excel_path)


def generate_report(
    profile: dict, opportunities: list[dict], role: str, location: str, output_path: str | Path | None = None
) -> str:
    """
    Compile analysis results into both a markdown report and Excel spreadsheet.

    Returns:
        Absolute path to the saved markdown report file.
    """
    now = datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H%M%S")
    _TMP_DIR.mkdir(parents=True, exist_ok=True)

    sorted_opps = sorted(
        opportunities,
        key=lambda x: x.get("match_score", 0) if isinstance(x.get("match_score"), (int, float)) else 0,
        reverse=True,
    )

    # --- Generate Markdown ---
    md_filename = f"opportunity_report_{timestamp}.md"
    report_path = Path(output_path) if output_path else _TMP_DIR / md_filename
    report_path.parent.mkdir(parents=True, exist_ok=True)

    sections = [
        "# AI Opportunity Radar Report",
        "",
        f"**Generated:** {now.strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "---",
        "",
        "## Profile Summary",
        "",
        _profile_summary_md(profile),
        "",
        "## Search Parameters",
        "",
        f"- **Target Role:** {role}",
        f"- **Location(s):** {location}",
        "",
        "## Opportunities Overview",
        "",
        _opportunities_table_md(sorted_opps),
        "",
        "## Detailed Opportunity Analysis",
        "",
    ]

    for i, opp in enumerate(sorted_opps, 1):
        sections.append(_opportunity_detail_md(opp, i))

    sections.append(_summary_stats_md(sorted_opps, profile))

    report_content = "\n".join(sections)
    report_path.write_text(report_content, encoding="utf-8")
    logger.info("Markdown report saved to %s", report_path)

    # --- Generate Excel ---
    excel_path = _TMP_DIR / f"opportunity_report_{timestamp}.xlsx"
    _generate_excel(profile, sorted_opps, role, location, excel_path)

    return str(report_path)


def main() -> None:
    """CLI entry point for testing and workflow integration."""
    import argparse

    parser = argparse.ArgumentParser(description="Generate AI Opportunity Radar markdown report")
    parser.add_argument("--input", required=True, help="Path to scored opportunities JSON")
    parser.add_argument("--profile", required=True, help="Path to profile JSON")
    parser.add_argument("--role", required=True, help="Target role used in search")
    parser.add_argument("--location", required=True, help="Location(s) used in search")
    parser.add_argument("--output", help="Override output path (default: .tmp/opportunity_report_YYYY-MM-DD_HHmmss.md)")
    args = parser.parse_args()

    with open(args.input, encoding="utf-8") as f:
        data = json.load(f)

    with open(args.profile, encoding="utf-8") as f:
        profile = json.load(f)

    opportunities = data.get("opportunities", data.get("results", data))
    if isinstance(opportunities, dict):
        opportunities = opportunities.get("opportunities", opportunities.get("results", []))
    if not isinstance(opportunities, list):
        opportunities = []

    path = generate_report(profile, opportunities, args.role, args.location, output_path=args.output)
    print(path)


if __name__ == "__main__":
    main()
